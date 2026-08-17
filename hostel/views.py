from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.core.paginator import Paginator

from django.db.models import Sum, Count, Q, F
from django.db.models.functions import ExtractMonth

from django.utils import timezone

from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash

from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet

import openpyxl
from openpyxl import Workbook

from .models import (
    Room,
    Bed,
    Student,
    Rent,
    Food,
    Expense
)

from .forms import (
    RoomForm,
    StudentForm,
    RentForm,
    FoodForm,
    ExpenseForm
)


# ==========================================================
# Dashboard
# ==========================================================

@login_required
def dashboard(request):

    # ------------------------------------------------------
    # Dashboard Counts
    # ------------------------------------------------------

    total_students = Student.objects.count()

    total_rooms = Room.objects.count()

    total_beds = Bed.objects.count()

    occupied_beds = Bed.objects.filter(
        status="Occupied"
    ).count()

    vacant_beds = Bed.objects.filter(
        status="Vacant"
    ).count()

    # ------------------------------------------------------
    # Dashboard Alerts
    # ------------------------------------------------------

    active_students = Student.objects.filter(
        status="Active"
    ).count()

    pending_rent_count = Rent.objects.filter(
        status="Pending"
    ).count()

    pending_rent_amount = Rent.objects.filter(
        status="Pending"
    ).aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    total_expenses = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    # ------------------------------------------------------
    # Low Bed Alert
    # ------------------------------------------------------

    low_bed_alert = vacant_beds <= 2

    # ------------------------------------------------------
    # Occupancy Percentage
    # ------------------------------------------------------

    if total_beds > 0:

        occupancy_percentage = round(
            (occupied_beds / total_beds) * 100,
            1
        )

    else:

        occupancy_percentage = 0

    # ------------------------------------------------------
    # Financial Summary
    # ------------------------------------------------------

    total_income = Rent.objects.aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    total_expense = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    pending_rent = Rent.objects.aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    net_profit = total_income - total_expense

    # ------------------------------------------------------
    # Today's Income & Expense
    # ------------------------------------------------------

    today = timezone.now().date()

    today_income = Rent.objects.filter(
        payment_date=today
    ).aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    today_expense = Expense.objects.filter(
        expense_date=today
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0

    today_profit = today_income - today_expense

    # ------------------------------------------------------
    # Monthly Income
    # ------------------------------------------------------

    monthly_income = (
        Rent.objects
        .annotate(
            month_num=ExtractMonth("payment_date")
        )
        .values("month_num")
        .annotate(
            total=Sum("paid_amount")
        )
        .order_by("month_num")
    )

    # ------------------------------------------------------
    # Monthly Expense
    # ------------------------------------------------------

    monthly_expense = (
        Expense.objects
        .annotate(
            month_num=ExtractMonth("expense_date")
        )
        .values("month_num")
        .annotate(
            total=Sum("amount")
        )
        .order_by("month_num")
    )

    # ------------------------------------------------------
    # Recent Students
    # ------------------------------------------------------

    recent_students = (
        Student.objects
        .select_related("room", "bed")
        .order_by("-id")[:5]
    )

    # ------------------------------------------------------
    # Recent Rent Payments
    # ------------------------------------------------------

    recent_rents = (
        Rent.objects
        .select_related("student")
        .order_by("-id")[:5]
    )

    # ------------------------------------------------------
    # Recent Expenses
    # ------------------------------------------------------

    recent_expenses = (
        Expense.objects
        .order_by("-id")[:5]
    )

    # ------------------------------------------------------
    # Context
    # ------------------------------------------------------

    context = {

        "total_students": total_students,
        "total_rooms": total_rooms,
        "total_beds": total_beds,

        "occupancy_percentage": occupancy_percentage,

        "today_income": today_income,
        "today_expense": today_expense,
        "today_profit": today_profit,

        "occupied_beds": occupied_beds,
        "vacant_beds": vacant_beds,

        "active_students": active_students,

        "pending_rent_count": pending_rent_count,
        "pending_rent_amount": pending_rent_amount,

        "total_expenses": total_expenses,

        "low_bed_alert": low_bed_alert,

        "pending_rent": pending_rent,
        "net_profit": net_profit,

        "monthly_income": list(monthly_income),
        "monthly_expense": list(monthly_expense),

        "recent_students": recent_students,
        "recent_rents": recent_rents,
        "recent_expenses": recent_expenses,

    }

    return render(
        request,
        "dashboard.html",
        context
    )


# ==========================================================
# Authentication
# ==========================================================

def login_view(request):

    if request.user.is_authenticated:

        return redirect("dashboard")

    if request.method == "POST":

        username = request.POST.get("username")

        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None:

            login(request, user)

            return redirect("dashboard")

        messages.error(
            request,
            "Invalid Username or Password"
        )

    return render(
        request,
        "registration/login.html"
    )


@login_required
def logout_view(request):

    logout(request)

    return redirect("login")


# ==========================================================
# Room Management
# ==========================================================

@login_required
def room_list(request):

    search = request.GET.get("search")

    room_type = request.GET.get("room_type")

    sharing_type = request.GET.get("sharing_type")

    rooms = Room.objects.all().order_by(
        "room_number"
    )

    if search:

        rooms = rooms.filter(
            room_number__icontains=search
        )

    if room_type:

        rooms = rooms.filter(
            room_type=room_type
        )

    if sharing_type:

        rooms = rooms.filter(
            sharing_type=sharing_type
        )

    paginator = Paginator(
        rooms,
        10
    )

    page_number = request.GET.get("page")

    rooms = paginator.get_page(
        page_number
    )

    total_rooms = Room.objects.count()

    total_beds = Bed.objects.count()

    occupied_beds = Bed.objects.filter(
        status="Occupied"
    ).count()

    vacant_beds = Bed.objects.filter(
        status="Vacant"
    ).count()

    for room in rooms:

        total = room.total_beds or 0

        available = room.available_beds or 0

        occupied = total - available

        room.occupied = occupied

        room.occupancy_percent = (
            int((occupied / total) * 100)
            if total
            else 0
        )

    return render(
        request,
        "rooms/room_list.html",
        {
            "rooms": rooms,

            "search": search,

            "selected_room_type": room_type,

            "selected_sharing_type": sharing_type,

            "total_rooms": total_rooms,

            "total_beds": total_beds,

            "occupied_beds": occupied_beds,

            "vacant_beds": vacant_beds,
        }
    )


@login_required
def add_room(request):

    if request.method == "POST":

        form = RoomForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Room added successfully."
            )

            return redirect(
                "room_list"
            )

    else:

        form = RoomForm()

    return render(
        request,
        "rooms/add_room.html",
        {
            "form": form
        }
    )


@login_required
def edit_room(request, pk):

    room = get_object_or_404(
        Room,
        pk=pk
    )

    if request.method == "POST":

        form = RoomForm(
            request.POST,
            instance=room
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Room updated successfully."
            )

            return redirect(
                "room_list"
            )

    else:

        form = RoomForm(
            instance=room
        )

    return render(
        request,
        "rooms/edit_room.html",
        {
            "form": form
        }
    )


@login_required
def delete_room(request, pk):

    room = get_object_or_404(
        Room,
        pk=pk
    )

    if request.method != "POST":

        messages.error(
            request,
            "Invalid request."
        )

        return redirect(
            "room_list"
        )

    if Student.objects.filter(
        room=room
    ).exists():

        messages.error(
            request,
            "Cannot delete this room because students are assigned to it."
        )

        return redirect(
            "room_list"
        )

    room.delete()

    messages.success(
        request,
        "Room deleted successfully."
    )

    return redirect(
        "room_list"
    )


# ==========================================================
# Bed Management
# ==========================================================

@login_required
def bed_list(request):

    search = request.GET.get("search")

    room = request.GET.get("room")

    status = request.GET.get("status")

    beds = Bed.objects.select_related(
        "room"
    )

    if search:

        beds = beds.filter(
            bed_number__icontains=search
        )

    if room:

        beds = beds.filter(
            room_id=room
        )

    if status:

        beds = beds.filter(
            status=status
        )

    # ------------------------------------------------------
    # Status Color
    # ------------------------------------------------------

    for bed in beds:

        if bed.status == "Occupied":

            bed.status_color = "danger"

        else:

            bed.status_color = "success"

    # ------------------------------------------------------
    # Pagination
    # ------------------------------------------------------

    paginator = Paginator(
        beds,
        10
    )

    page_number = request.GET.get("page")

    beds = paginator.get_page(
        page_number
    )

    total_beds = Bed.objects.count()

    occupied_beds = Bed.objects.filter(
        status="Occupied"
    ).count()

    vacant_beds = Bed.objects.filter(
        status="Vacant"
    ).count()

    total_rooms = Room.objects.count()

    return render(
        request,
        "beds/bed_list.html",
        {
            "beds": beds,

            "rooms": Room.objects.all(),

            "total_beds": total_beds,

            "occupied_beds": occupied_beds,

            "vacant_beds": vacant_beds,

            "total_rooms": total_rooms,

            "selected_room": room,

            "selected_status": status,

            "search": search,
        }
    )


# ==========================================================
# Student Management
# ==========================================================

@login_required
def student_list(request):

    search = request.GET.get(
        "search"
    )

    sort = request.GET.get(
        "sort",
        "name"
    )

    room = request.GET.get(
        "room"
    )

    # ------------------------------------------------------
    # Base Query
    # ------------------------------------------------------

    students = Student.objects.select_related(
        "room",
        "bed"
    )

    # ------------------------------------------------------
    # Search
    # ------------------------------------------------------

    if search:

        students = students.filter(
            name__icontains=search
        )

    # ------------------------------------------------------
    # Room Filter
    # ------------------------------------------------------

    if room:

        students = students.filter(
            room_id=room
        )

    # ------------------------------------------------------
    # Sorting
    # ------------------------------------------------------

    if sort == "name":

        students = students.order_by(
            "name"
        )

    elif sort == "-name":

        students = students.order_by(
            "-name"
        )

    elif sort == "room":

        students = students.order_by(
            "room__room_number"
        )

    elif sort == "latest":

        students = students.order_by(
            "-id"
        )

    else:

        students = students.order_by(
            "name"
        )

    # ------------------------------------------------------
    # Pagination
    # ------------------------------------------------------

    paginator = Paginator(
        students,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    students = paginator.get_page(
        page_number
    )

    # ------------------------------------------------------
    # Statistics
    # ------------------------------------------------------

    total_students = Student.objects.count()

    active_students = Student.objects.filter(
        status="Active"
    ).count()

    occupied_rooms = Room.objects.filter(
        available_beds__lt=F("total_beds")
    ).count()

    vacant_beds = Bed.objects.filter(
        status="Vacant"
    ).count()

    # ------------------------------------------------------
    # Render
    # ------------------------------------------------------

    return render(
        request,
        "students/student_list.html",
        {
            "students": students,

            "rooms": Room.objects.all(),

            "sort": sort,

            "selected_room": room,

            "search": search,

            "total_students": total_students,

            "active_students": active_students,

            "occupied_rooms": occupied_rooms,

            "vacant_beds": vacant_beds,
        }
    )


@login_required
def student_detail(request, student_id):

    student = get_object_or_404(
        Student,
        id=student_id
    )

    rent_history = Rent.objects.filter(
        student=student
    ).order_by(
        "-payment_date"
    )

    total_paid = sum(
        rent.paid_amount or 0
        for rent in rent_history
    )

    total_pending = sum(
        rent.pending_amount or 0
        for rent in rent_history
    )

    return render(
        request,
        "students/student_detail.html",
        {
            "student": student,

            "rent_history": rent_history,

            "total_paid": total_paid,

            "total_pending": total_pending,
        }
    )


@login_required
def add_student(request):

    if request.method == "POST":

        form = StudentForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Student added successfully."
            )

            return redirect(
                "student_list"
            )

    else:

        form = StudentForm()

    return render(
        request,
        "students/add_student.html",
        {
            "form": form
        }
    )


@login_required
def edit_student(request, pk):

    student = get_object_or_404(
        Student,
        pk=pk
    )

    if request.method == "POST":

        form = StudentForm(
            request.POST,
            request.FILES,
            instance=student
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Student updated successfully."
            )

            return redirect(
                "student_list"
            )

    else:

        form = StudentForm(
            instance=student
        )

    return render(
        request,
        "students/edit_student.html",
        {
            "form": form
        }
    )


@login_required
def delete_student(request, pk):

    student = get_object_or_404(
        Student,
        pk=pk
    )

    bed = student.bed

    room = student.room

    # ------------------------------------------------------
    # Vacate Bed
    # ------------------------------------------------------

    if bed:

        bed.status = "Vacant"

        bed.save()

    # ------------------------------------------------------
    # Delete Student
    # ------------------------------------------------------

    student.delete()

    # ------------------------------------------------------
    # Recalculate Room Beds
    # ------------------------------------------------------

    if room:

        room.available_beds = room.beds.filter(
            status="Vacant"
        ).count()

        room.save()

    messages.success(
        request,
        "Student deleted successfully."
    )

    return redirect(
        "student_list"
    )


# ==========================================================
# Rent Management
# ==========================================================

@login_required
def rent_list(request):

    search = request.GET.get(
        "search"
    )

    rents = Rent.objects.select_related(
        "student"
    ).order_by(
        "-id"
    )

    if search:

        rents = rents.filter(
            student__name__icontains=search
        )

    paginator = Paginator(
        rents,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    rents = paginator.get_page(
        page_number
    )

    return render(
        request,
        "rent/rent_list.html",
        {
            "rents": rents
        }
    )


@login_required
def add_rent(request):

    if request.method == "POST":

        form = RentForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Rent Payment Added Successfully."
            )

            return redirect(
                "rent_list"
            )

        messages.error(
            request,
            "Please correct the errors below."
        )

    else:

        form = RentForm()

    return render(
        request,
        "rent/add_rent.html",
        {
            "form": form
        }
    )


@login_required
def edit_rent(request, pk):

    rent = get_object_or_404(
        Rent,
        pk=pk
    )

    if request.method == "POST":

        form = RentForm(
            request.POST,
            instance=rent
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Rent payment updated successfully."
            )

            return redirect(
                "rent_list"
            )

    else:

        form = RentForm(
            instance=rent
        )

    return render(
        request,
        "rent/edit_rent.html",
        {
            "form": form
        }
    )


@login_required
def delete_rent(request, pk):

    rent = get_object_or_404(
        Rent,
        pk=pk
    )

    rent.delete()

    messages.success(
        request,
        "Rent payment deleted successfully."
    )

    return redirect(
        "rent_list"
    )


@login_required
def payment_receipt(request, pk):

    rent = get_object_or_404(
        Rent,
        pk=pk
    )

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="Receipt_{rent.id}.pdf"'
    )

    p = canvas.Canvas(
        response
    )

    # ------------------------------------------------------
    # Header
    # ------------------------------------------------------

    p.setFont(
        "Helvetica-Bold",
        20
    )

    p.drawCentredString(
        300,
        800,
        "SSMR MENS PG"
    )

    p.setFont(
        "Helvetica",
        12
    )

    p.drawCentredString(
        300,
        780,
        "Hostel Rent Payment Receipt"
    )

    p.line(
        40,
        770,
        560,
        770
    )

    y = 735

    p.drawString(
        50,
        y,
        f"Receipt No : {rent.id}"
    )

    y -= 25

    p.drawString(
        50,
        y,
        f"Student : {rent.student.name}"
    )

    y -= 25

    room_number = "-"

    if rent.student.room:

        room_number = rent.student.room.room_number

    p.drawString(
        50,
        y,
        f"Room : {room_number}"
    )

    y -= 25

    bed_number = "-"

    if rent.student.bed:

        bed_number = rent.student.bed.bed_number

    p.drawString(
        50,
        y,
        f"Bed : {bed_number}"
    )

    y -= 25

    p.drawString(
        50,
        y,
        f"Month : {rent.month}"
    )

    y -= 25

    p.drawString(
        50,
        y,
        f"Year : {rent.year}"
    )

    y -= 25

    p.drawString(
        50,
        y,
        f"Paid Amount : Rs. {rent.paid_amount}"
    )

    y -= 25

    p.drawString(
        50,
        y,
        f"Pending Amount : Rs. {rent.pending_amount}"
    )

    y -= 25

    p.drawString(
        50,
        y,
        f"Payment Date : {rent.payment_date}"
    )

    y -= 60

    p.line(
        350,
        y,
        520,
        y
    )

    p.drawString(
        380,
        y - 20,
        "Authorized Signature"
    )

    p.save()

    return response


# ==========================================================
# Food Management
# ==========================================================

@login_required
def food_list(request):

    search = request.GET.get(
        "search"
    )

    foods = Food.objects.select_related(
        "student"
    ).order_by(
        "-id"
    )

    if search:

        foods = foods.filter(
            student__name__icontains=search
        )

    paginator = Paginator(
        foods,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    foods = paginator.get_page(
        page_number
    )

    return render(
        request,
        "food/food_list.html",
        {
            "foods": foods
        }
    )


@login_required
def add_food(request):

    if request.method == "POST":

        form = FoodForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Food record added successfully."
            )

            return redirect(
                "food_list"
            )

    else:

        form = FoodForm()

    return render(
        request,
        "food/add_food.html",
        {
            "form": form
        }
    )


@login_required
def edit_food(request, pk):

    food = get_object_or_404(
        Food,
        pk=pk
    )

    if request.method == "POST":

        form = FoodForm(
            request.POST,
            instance=food
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Food record updated successfully."
            )

            return redirect(
                "food_list"
            )

    else:

        form = FoodForm(
            instance=food
        )

    return render(
        request,
        "food/edit_food.html",
        {
            "form": form
        }
    )


@login_required
def delete_food(request, pk):

    food = get_object_or_404(
        Food,
        pk=pk
    )

    food.delete()

    messages.success(
        request,
        "Food record deleted successfully."
    )

    return redirect(
        "food_list"
    )


@login_required
def food_report(request):

    foods = Food.objects.all()

    breakfast = foods.filter(
        breakfast=True
    ).count()

    lunch = foods.filter(
        lunch=True
    ).count()

    dinner = foods.filter(
        dinner=True
    ).count()

    return render(
        request,
        "food/food_report.html",
        {
            "breakfast": breakfast,

            "lunch": lunch,

            "dinner": dinner,
        }
    )


# ==========================================================
# Expense Management
# ==========================================================

@login_required
def expense_list(request):

    search = request.GET.get(
        "search"
    )

    expenses = Expense.objects.all().order_by(
        "-expense_date"
    )

    if search:

        expenses = expenses.filter(
            category__icontains=search
        )

    paginator = Paginator(
        expenses,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    expenses = paginator.get_page(
        page_number
    )

    return render(
        request,
        "expense/expense_list.html",
        {
            "expenses": expenses
        }
    )


@login_required
def add_expense(request):

    if request.method == "POST":

        form = ExpenseForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Expense added successfully."
            )

            return redirect(
                "expense_list"
            )

    else:

        form = ExpenseForm()

    return render(
        request,
        "expense/add_expense.html",
        {
            "form": form
        }
    )


@login_required
def edit_expense(request, pk):

    expense = get_object_or_404(
        Expense,
        pk=pk
    )

    if request.method == "POST":

        form = ExpenseForm(
            request.POST,
            instance=expense
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Expense updated successfully."
            )

            return redirect(
                "expense_list"
            )

    else:

        form = ExpenseForm(
            instance=expense
        )

    return render(
        request,
        "expense/edit_expense.html",
        {
            "form": form
        }
    )


@login_required
def delete_expense(request, pk):

    expense = get_object_or_404(
        Expense,
        pk=pk
    )

    expense.delete()

    messages.success(
        request,
        "Expense deleted successfully."
    )

    return redirect(
        "expense_list"
    )


# ==========================================================
# Student Statistics
# ==========================================================

@login_required
def student_statistics(request):

    total_students = Student.objects.count()

    active_students = Student.objects.filter(
        status="Active"
    ).count()

    left_students = Student.objects.filter(
        status="Left"
    ).count()

    veg_students = Student.objects.filter(
        food_type="Veg"
    ).count()

    nonveg_students = Student.objects.filter(
        food_type="Non Veg"
    ).count()

    college_data = (
        Student.objects
        .values("college")
        .annotate(
            total=Count("id"),
            active=Count(
                "id",
                filter=Q(
                    status="Active"
                )
            ),
            left=Count(
                "id",
                filter=Q(
                    status="Left"
                )
            ),
        )
        .order_by("-total")
    )

    monthly_joinings = (
        Student.objects
        .annotate(
            month=ExtractMonth(
                "joining_date"
            )
        )
        .values("month")
        .annotate(
            total=Count("id")
        )
        .order_by("month")
    )

    context = {

        "total_students": total_students,

        "active_students": active_students,

        "left_students": left_students,

        "monthly_joinings": monthly_joinings,

        "veg_students": veg_students,

        "nonveg_students": nonveg_students,

        "college_data": college_data,

    }

    return render(
        request,
        "reports/student_statistics.html",
        context
    )


# ==========================================================
# Student Statistics PDF
# ==========================================================

@login_required
def student_statistics_pdf(request):

    students = Student.objects.select_related(
        "room",
        "bed"
    ).order_by(
        "name"
    )

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="student_statistics.pdf"'
    )

    p = canvas.Canvas(
        response
    )

    # ------------------------------------------------------
    # Title
    # ------------------------------------------------------

    p.setFont(
        "Helvetica-Bold",
        18
    )

    p.drawString(
        180,
        800,
        "SSMR MENS PG"
    )

    p.setFont(
        "Helvetica-Bold",
        14
    )

    p.drawString(
        170,
        775,
        "Student Statistics Report"
    )

    # ------------------------------------------------------
    # Summary
    # ------------------------------------------------------

    total_students = Student.objects.count()

    active_students = Student.objects.filter(
        status="Active"
    ).count()

    left_students = Student.objects.filter(
        status="Left"
    ).count()

    veg_students = Student.objects.filter(
        food_type="Veg"
    ).count()

    nonveg_students = Student.objects.filter(
        food_type="Non Veg"
    ).count()

    y = 740

    p.setFont(
        "Helvetica",
        10
    )

    p.drawString(
        50,
        y,
        f"Total Students: {total_students}"
    )

    p.drawString(
        200,
        y,
        f"Active: {active_students}"
    )

    p.drawString(
        330,
        y,
        f"Left: {left_students}"
    )

    p.drawString(
        430,
        y,
        f"Veg: {veg_students}"
    )

    p.drawString(
        500,
        y,
        f"Non Veg: {nonveg_students}"
    )

    # ------------------------------------------------------
    # Table Header
    # ------------------------------------------------------

    y -= 40

    p.setFont(
        "Helvetica-Bold",
        10
    )

    p.drawString(
        40,
        y,
        "S.No"
    )

    p.drawString(
        75,
        y,
        "Name"
    )

    p.drawString(
        220,
        y,
        "Phone"
    )

    p.drawString(
        310,
        y,
        "Room"
    )

    p.drawString(
        370,
        y,
        "Bed"
    )

    p.drawString(
        430,
        y,
        "Food"
    )

    p.drawString(
        490,
        y,
        "Status"
    )

    y -= 20

    p.setFont(
        "Helvetica",
        9
    )

    # ------------------------------------------------------
    # Student Data
    # ------------------------------------------------------

    for index, student in enumerate(
        students,
        start=1
    ):

        if y < 50:

            p.showPage()

            y = 800

            p.setFont(
                "Helvetica",
                9
            )

        room_number = "-"

        if student.room:

            room_number = str(
                student.room.room_number
            )

        bed_number = "-"

        if student.bed:

            bed_number = str(
                student.bed.bed_number
            )

        p.drawString(
            40,
            y,
            str(index)
        )

        p.drawString(
            75,
            y,
            str(student.name)[:22]
        )

        p.drawString(
            220,
            y,
            str(student.phone or "")[:15]
        )

        p.drawString(
            310,
            y,
            room_number
        )

        p.drawString(
            370,
            y,
            bed_number
        )

        p.drawString(
            430,
            y,
            str(student.food_type or "")
        )

        p.drawString(
            490,
            y,
            str(student.status or "")
        )

        y -= 20

    # ------------------------------------------------------
    # Footer
    # ------------------------------------------------------

    p.setFont(
        "Helvetica",
        8
    )

    p.drawString(
        40,
        25,
        "SSMR MENS PG - Hostel Management System"
    )

    p.save()

    return response


# ==========================================================
# Student Statistics Excel
# ==========================================================

@login_required
def student_statistics_excel(request):

    students = Student.objects.select_related(
        "room",
        "bed"
    ).order_by(
        "name"
    )

    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    worksheet.title = "Student Statistics"

    # ------------------------------------------------------
    # Title
    # ------------------------------------------------------

    worksheet["A1"] = "SSMR MENS PG"

    worksheet["A2"] = "Student Statistics Report"

    # ------------------------------------------------------
    # Summary
    # ------------------------------------------------------

    worksheet["A4"] = "Total Students"

    worksheet["B4"] = Student.objects.count()

    worksheet["C4"] = "Active Students"

    worksheet["D4"] = Student.objects.filter(
        status="Active"
    ).count()

    worksheet["E4"] = "Left Students"

    worksheet["F4"] = Student.objects.filter(
        status="Left"
    ).count()

    worksheet["G4"] = "Veg Students"

    worksheet["H4"] = Student.objects.filter(
        food_type="Veg"
    ).count()

    worksheet["I4"] = "Non Veg Students"

    worksheet["J4"] = Student.objects.filter(
        food_type="Non Veg"
    ).count()

    # ------------------------------------------------------
    # Table Header
    # ------------------------------------------------------

    headers = [

        "S.No",
        "Name",
        "Phone",
        "Aadhaar",
        "College / Workplace",
        "Room",
        "Bed",
        "Food Type",
        "Joining Date",
        "Deposit",
        "Status",

    ]

    for column, header in enumerate(
        headers,
        start=1
    ):

        worksheet.cell(
            row=6,
            column=column
        ).value = header

    # ------------------------------------------------------
    # Student Data
    # ------------------------------------------------------

    row = 7

    for index, student in enumerate(
        students,
        start=1
    ):

        room_number = ""

        if student.room:

            room_number = (
                student.room.room_number
            )

        bed_number = ""

        if student.bed:

            bed_number = (
                student.bed.bed_number
            )

        deposit = 0

        if student.deposit is not None:

            deposit = float(
                student.deposit
            )

        worksheet.cell(
            row=row,
            column=1
        ).value = index

        worksheet.cell(
            row=row,
            column=2
        ).value = student.name

        worksheet.cell(
            row=row,
            column=3
        ).value = student.phone

        worksheet.cell(
            row=row,
            column=4
        ).value = student.aadhaar

        worksheet.cell(
            row=row,
            column=5
        ).value = student.college

        worksheet.cell(
            row=row,
            column=6
        ).value = room_number

        worksheet.cell(
            row=row,
            column=7
        ).value = bed_number

        worksheet.cell(
            row=row,
            column=8
        ).value = student.food_type

        worksheet.cell(
            row=row,
            column=9
        ).value = student.joining_date

        worksheet.cell(
            row=row,
            column=10
        ).value = deposit

        worksheet.cell(
            row=row,
            column=11
        ).value = student.status

        row += 1

    # ------------------------------------------------------
    # Column Width
    # ------------------------------------------------------

    column_widths = {

        "A": 8,
        "B": 25,
        "C": 15,
        "D": 16,
        "E": 30,
        "F": 12,
        "G": 12,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 15,

    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    worksheet.freeze_panes = "A7"

    # ------------------------------------------------------
    # Response
    # ------------------------------------------------------

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="student_statistics.xlsx"'
    )

    workbook.save(response)

    return response


# ==========================================================
# Student Statistics Print
# ==========================================================

@login_required
def student_statistics_print(request):

    students = Student.objects.select_related(
        "room",
        "bed"
    ).order_by(
        "name"
    )

    total_students = Student.objects.count()

    active_students = Student.objects.filter(
        status="Active"
    ).count()

    left_students = Student.objects.filter(
        status="Left"
    ).count()

    veg_students = Student.objects.filter(
        food_type="Veg"
    ).count()

    nonveg_students = Student.objects.filter(
        food_type="Non Veg"
    ).count()

    return render(
        request,
        "reports/student_statistics_print.html",
        {
            "students": students,

            "total_students": total_students,

            "active_students": active_students,

            "left_students": left_students,

            "veg_students": veg_students,

            "nonveg_students": nonveg_students,
        }
    )


# ==========================================================
# Financial Analytics
# ==========================================================

@login_required
def financial_analytics(request):

    # ------------------------------------------------------
    # Rent Statistics
    # ------------------------------------------------------

    total_rent = Rent.objects.filter(
        status="Paid"
    ).aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    pending_rent = Rent.objects.filter(
        status="Pending"
    ).aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    # ------------------------------------------------------
    # Expense Statistics
    # ------------------------------------------------------

    total_expenses = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    # ------------------------------------------------------
    # Net Income
    # ------------------------------------------------------

    net_income = (
        total_rent -
        total_expenses
    )

    # ------------------------------------------------------
    # Monthly Income
    # ------------------------------------------------------

    monthly_income = (
        Rent.objects
        .filter(status="Paid")
        .annotate(
            payment_month=ExtractMonth(
                "payment_date"
            )
        )
        .values(
            "payment_month"
        )
        .annotate(
            total=Sum("paid_amount")
        )
        .order_by(
            "payment_month"
        )
    )

    # ------------------------------------------------------
    # Monthly Expenses
    # ------------------------------------------------------

    monthly_expense = (
        Expense.objects
        .annotate(
            expense_month=ExtractMonth(
                "expense_date"
            )
        )
        .values(
            "expense_month"
        )
        .annotate(
            total=Sum("amount")
        )
        .order_by(
            "expense_month"
        )
    )

    context = {

        "total_rent": total_rent,

        "pending_rent": pending_rent,

        "total_expenses": total_expenses,

        "net_income": net_income,

        "monthly_income": monthly_income,

        "monthly_expense": monthly_expense,

    }

    return render(
        request,
        "reports/financial_analytics.html",
        context
    )


# ==========================================================
# Financial Analytics PDF
# ==========================================================

@login_required
def financial_analytics_pdf(request):

    total_rent = Rent.objects.filter(
        status="Paid"
    ).aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    pending_rent = Rent.objects.filter(
        status="Pending"
    ).aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    total_expenses = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    net_income = (
        total_rent -
        total_expenses
    )

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="financial_analytics.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    elements = []

    # ------------------------------------------------------
    # Header
    # ------------------------------------------------------

    elements.append(
        Paragraph(
            "SSMR MENS PG",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "Financial Analytics Report",
            styles["Heading2"]
        )
    )

    elements.append(
        Spacer(
            1,
            15
        )
    )

    # ------------------------------------------------------
    # Financial Summary
    # ------------------------------------------------------

    summary_data = [

        [
            "Financial Summary",
            "Amount"
        ],

        [
            "Rent Collected",
            f"Rs. {total_rent}"
        ],

        [
            "Pending Rent",
            f"Rs. {pending_rent}"
        ],

        [
            "Total Expenses",
            f"Rs. {total_expenses}"
        ],

        [
            "Net Income",
            f"Rs. {net_income}"
        ],

    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            300,
            180
        ],
        repeatRows=1
    )

    summary_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkblue
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "ALIGN",
                (1, 0),
                (1, -1),
                "RIGHT"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.whitesmoke
                ]
            ),

            (
                "BACKGROUND",
                (0, 4),
                (-1, 4),
                colors.lightgreen
            ),

        ])
    )

    elements.append(
        summary_table
    )

    elements.append(
        Spacer(
            1,
            20
        )
    )

    # ------------------------------------------------------
    # Expense Heading
    # ------------------------------------------------------

    elements.append(
        Paragraph(
            "Expense Records",
            styles["Heading3"]
        )
    )

    elements.append(
        Spacer(
            1,
            10
        )
    )

    # ------------------------------------------------------
    # Expense Table
    # ------------------------------------------------------

    expense_data = [[

        "Date",
        "Category",
        "Description",
        "Paid To",
        "Mode",
        "Amount",
        "Receipt",

    ]]

    expenses = Expense.objects.order_by(
        "-expense_date"
    )

    for expense in expenses:

        expense_date = (
            str(expense.expense_date)
            if expense.expense_date
            else "-"
        )

        category = (
            str(expense.category)
            if expense.category
            else "-"
        )

        description = (
            str(expense.description)
            if expense.description
            else "-"
        )

        paid_to = (
            str(expense.paid_to)
            if expense.paid_to
            else "-"
        )

        payment_mode = (
            str(expense.payment_mode)
            if expense.payment_mode
            else "-"
        )

        amount = (
            f"Rs. {expense.amount}"
            if expense.amount is not None
            else "Rs. 0"
        )

        receipt = (
            str(expense.receipt_number)
            if expense.receipt_number
            else "-"
        )

        expense_data.append([

            expense_date,

            category[:18],

            description[:25],

            paid_to[:18],

            payment_mode[:12],

            amount,

            receipt[:15],

        ])

    expense_table = Table(
        expense_data,
        colWidths=[
            65,
            75,
            110,
            75,
            55,
            65,
            70,
        ],
        repeatRows=1
    )

    expense_table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkred
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                7
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.whitesmoke
                ]
            ),

        ])
    )

    elements.append(
        expense_table
    )

    elements.append(
        Spacer(
            1,
            15
        )
    )

    elements.append(
        Paragraph(
            "SSMR MENS PG - Hostel Management System",
            styles["Normal"]
        )
    )

    doc.build(
        elements
    )

    return response


# ==========================================================
# Financial Analytics Excel
# ==========================================================

@login_required
def financial_analytics_excel(request):

    total_rent = Rent.objects.filter(
        status="Paid"
    ).aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    pending_rent = Rent.objects.filter(
        status="Pending"
    ).aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    total_expenses = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    net_income = (
        total_rent -
        total_expenses
    )

    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    worksheet.title = (
        "Financial Analytics"
    )

    # ------------------------------------------------------
    # Title
    # ------------------------------------------------------

    worksheet["A1"] = "SSMR MENS PG"

    worksheet["A2"] = (
        "Financial Analytics Report"
    )

    # ------------------------------------------------------
    # Summary
    # ------------------------------------------------------

    worksheet["A4"] = (
        "Rent Collected"
    )

    worksheet["B4"] = float(
        total_rent
    )

    worksheet["C4"] = (
        "Pending Rent"
    )

    worksheet["D4"] = float(
        pending_rent
    )

    worksheet["E4"] = (
        "Total Expenses"
    )

    worksheet["F4"] = float(
        total_expenses
    )

    worksheet["G4"] = (
        "Net Income"
    )

    worksheet["H4"] = float(
        net_income
    )

    # ------------------------------------------------------
    # Expense Header
    # ------------------------------------------------------

    headers = [

        "S.No",
        "Date",
        "Category",
        "Description",
        "Paid To",
        "Payment Mode",
        "Amount",
        "Receipt Number",

    ]

    for column, header in enumerate(
        headers,
        start=1
    ):

        worksheet.cell(
            row=6,
            column=column
        ).value = header

    # ------------------------------------------------------
    # Expense Data
    # ------------------------------------------------------

    expenses = Expense.objects.order_by(
        "-expense_date"
    )

    row = 7

    for index, expense in enumerate(
        expenses,
        start=1
    ):

        amount = 0

        if expense.amount is not None:

            amount = float(
                expense.amount
            )

        worksheet.cell(
            row=row,
            column=1
        ).value = index

        worksheet.cell(
            row=row,
            column=2
        ).value = expense.expense_date

        worksheet.cell(
            row=row,
            column=3
        ).value = expense.category

        worksheet.cell(
            row=row,
            column=4
        ).value = expense.description

        worksheet.cell(
            row=row,
            column=5
        ).value = expense.paid_to

        worksheet.cell(
            row=row,
            column=6
        ).value = expense.payment_mode

        worksheet.cell(
            row=row,
            column=7
        ).value = amount

        worksheet.cell(
            row=row,
            column=8
        ).value = expense.receipt_number

        row += 1

    # ------------------------------------------------------
    # Column Widths
    # ------------------------------------------------------

    widths = {

        "A": 8,
        "B": 15,
        "C": 20,
        "D": 35,
        "E": 20,
        "F": 18,
        "G": 15,
        "H": 20,

    }

    for column, width in widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    worksheet.freeze_panes = "A7"

    # ------------------------------------------------------
    # Response
    # ------------------------------------------------------

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="financial_analytics.xlsx"'
    )

    workbook.save(
        response
    )

    return response


# ==========================================================
# Financial Analytics Print
# ==========================================================

@login_required
def financial_analytics_print(request):

    total_rent = Rent.objects.filter(
        status="Paid"
    ).aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    pending_rent = Rent.objects.filter(
        status="Pending"
    ).aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    total_expenses = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    net_income = (
        total_rent -
        total_expenses
    )

    expenses = Expense.objects.order_by(
        "-expense_date"
    )

    return render(
        request,
        "reports/financial_analytics_print.html",
        {
            "total_rent": total_rent,

            "pending_rent": pending_rent,

            "total_expenses": total_expenses,

            "net_income": net_income,

            "expenses": expenses,
        }
    )


# ==========================================================
# Reports Dashboard
# ==========================================================

@login_required
def reports_dashboard(request):

    total_rooms = Room.objects.count()

    total_beds = Bed.objects.count()

    occupied_beds = Bed.objects.filter(
        status="Occupied"
    ).count()

    vacant_beds = Bed.objects.filter(
        status="Vacant"
    ).count()

    total_students = Student.objects.count()

    total_income = Rent.objects.aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    pending_rent = Rent.objects.aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    total_expenses = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    occupancy_percentage = (

        round(
            (
                occupied_beds /
                total_beds
            ) * 100
        )

        if total_beds > 0

        else 0

    )

    return render(
        request,
        "reports/dashboard.html",
        {
            "total_rooms": total_rooms,

            "total_beds": total_beds,

            "occupied_beds": occupied_beds,

            "vacant_beds": vacant_beds,

            "total_students": total_students,

            "total_income": total_income,

            "pending_rent": pending_rent,

            "total_expense": total_expenses,

            "net_profit": (
                total_income -
                total_expenses
            ),

            "occupancy_percentage": (
                occupancy_percentage
            ),
        }
    )


# ==========================================================
# Student Report
# ==========================================================

@login_required
def student_report(request):

    search = request.GET.get(
        "search"
    )

    students = Student.objects.select_related(
        "room",
        "bed"
    ).all()

    if search:

        students = students.filter(
            name__icontains=search
        )

    paginator = Paginator(
        students,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    students = paginator.get_page(
        page_number
    )

    return render(
        request,
        "reports/student_report.html",
        {
            "students": students
        }
    )


# ==========================================================
# Room Report
# ==========================================================

@login_required
def room_report(request):

    rooms = Room.objects.all().order_by(
        "room_number"
    )

    for room in rooms:

        total_beds = room.total_beds or 0

        available_beds = (
            room.available_beds or 0
        )

        room.occupied = (
            total_beds -
            available_beds
        )

    total_rooms = Room.objects.count()

    total_beds = Bed.objects.count()

    occupied_beds = Bed.objects.filter(
        status="Occupied"
    ).count()

    vacant_beds = Bed.objects.filter(
        status="Vacant"
    ).count()

    context = {

        "rooms": rooms,

        "total_rooms": total_rooms,

        "total_beds": total_beds,

        "occupied_beds": occupied_beds,

        "vacant_beds": vacant_beds,

    }

    return render(
        request,
        "reports/room_report.html",
        context
    )


# ==========================================================
# Rent Report
# ==========================================================

@login_required
def rent_report(request):

    search = request.GET.get(
        "search"
    )

    rents = Rent.objects.select_related(
        "student"
    ).order_by(
        "-year",
        "month"
    )

    if search:

        rents = rents.filter(
            student__name__icontains=search
        )

    paginator = Paginator(
        rents,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    rents = paginator.get_page(
        page_number
    )

    total_collection = Rent.objects.aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    total_pending = Rent.objects.aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    monthly_analysis = (
        Rent.objects
        .values(
            "year",
            "month"
        )
        .annotate(
            collection=Sum(
                "paid_amount"
            ),
            pending=Sum(
                "pending_amount"
            ),
        )
        .order_by(
            "-year",
            "month"
        )
    )

    context = {

        "rents": rents,

        "total_collection": total_collection,

        "total_pending": total_pending,

        "monthly_analysis": monthly_analysis,

    }

    return render(
        request,
        "reports/rent_report.html",
        context
    )


# ==========================================================
# Expense Report
# ==========================================================

@login_required
def expense_report(request):

    search = request.GET.get(
        "search"
    )

    expenses = Expense.objects.all().order_by(
        "-expense_date"
    )

    if search:

        expenses = expenses.filter(
            category__icontains=search
        )

    paginator = Paginator(
        expenses,
        10
    )

    page_number = request.GET.get(
        "page"
    )

    expenses = paginator.get_page(
        page_number
    )

    total_expense = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    context = {

        "expenses": expenses,

        "total_expense": total_expense,

    }

    return render(
        request,
        "reports/expense_report.html",
        context
    )


# ==========================================================
# Finance Report
# ==========================================================

@login_required
def finance_report(request):

    total_income = Rent.objects.aggregate(
        total=Sum("paid_amount")
    )["total"] or 0

    total_expense = Expense.objects.aggregate(
        total=Sum("amount")
    )["total"] or 0

    pending_rent = Rent.objects.aggregate(
        total=Sum("pending_amount")
    )["total"] or 0

    net_profit = (
        total_income -
        total_expense
    )

    context = {

        "total_income": total_income,

        "total_expense": total_expense,

        "pending_rent": pending_rent,

        "net_profit": net_profit,

    }

    return render(
        request,
        "reports/finance_report.html",
        context
    )


# ==========================================================
# Student Report PDF
# ==========================================================

@login_required
def student_report_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="student_report.pdf"'
    )

    p = canvas.Canvas(
        response
    )

    # ------------------------------------------------------
    # Header
    # ------------------------------------------------------

    p.setFont(
        "Helvetica-Bold",
        18
    )

    p.drawCentredString(
        300,
        800,
        "SSMR MENS PG"
    )

    p.setFont(
        "Helvetica",
        12
    )

    p.drawCentredString(
        300,
        780,
        "Student Report"
    )

    p.line(
        40,
        770,
        560,
        770
    )

    students = Student.objects.select_related(
        "room",
        "bed"
    ).all()

    y = 740

    # ------------------------------------------------------
    # Header
    # ------------------------------------------------------

    p.setFont(
        "Helvetica-Bold",
        11
    )

    p.drawString(
        40,
        y,
        "Student"
    )

    p.drawString(
        200,
        y,
        "Room"
    )

    p.drawString(
        300,
        y,
        "Bed"
    )

    p.drawString(
        400,
        y,
        "Phone"
    )

    y -= 20

    p.setFont(
        "Helvetica",
        10
    )

    # ------------------------------------------------------
    # Records
    # ------------------------------------------------------

    for student in students:

        student_name = (
            str(student.name)
            if student.name
            else "-"
        )

        room_number = "-"

        if student.room:

            room_number = str(
                student.room.room_number
            )

        bed_number = "-"

        if student.bed:

            bed_number = str(
                student.bed.bed_number
            )

        phone = (
            str(student.phone)
            if student.phone
            else "-"
        )

        p.drawString(
            40,
            y,
            student_name[:25]
        )

        p.drawString(
            200,
            y,
            room_number[:15]
        )

        p.drawString(
            300,
            y,
            bed_number[:15]
        )

        p.drawString(
            400,
            y,
            phone[:20]
        )

        y -= 20

        if y < 50:

            p.showPage()

            y = 800

            p.setFont(
                "Helvetica-Bold",
                11
            )

            p.drawString(
                40,
                y,
                "Student"
            )

            p.drawString(
                200,
                y,
                "Room"
            )

            p.drawString(
                300,
                y,
                "Bed"
            )

            p.drawString(
                400,
                y,
                "Phone"
            )

            y -= 20

            p.setFont(
                "Helvetica",
                10
            )

    p.save()

    return response


# ==========================================================
# Student Report Excel
# ==========================================================

@login_required
def student_report_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Students"

    sheet.append([
        "Student Name",
        "Room",
        "Bed",
        "Phone"
    ])

    students = Student.objects.select_related(
        "room",
        "bed"
    ).all()

    for student in students:

        room_number = ""

        if student.room:

            room_number = (
                student.room.room_number
            )

        bed_number = ""

        if student.bed:

            bed_number = (
                student.bed.bed_number
            )

        sheet.append([
            student.name,
            room_number,
            bed_number,
            student.phone,
        ])

    sheet.column_dimensions[
        "A"
    ].width = 30

    sheet.column_dimensions[
        "B"
    ].width = 15

    sheet.column_dimensions[
        "C"
    ].width = 15

    sheet.column_dimensions[
        "D"
    ].width = 20

    sheet.freeze_panes = "A2"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="student_report.xlsx"'
    )

    workbook.save(
        response
    )

    return response


# ==========================================================
# Student Print Report
# ==========================================================

@login_required
def student_report_print(request):

    students = Student.objects.select_related(
        "room",
        "bed"
    ).all()

    return render(
        request,
        "reports/student_report_print.html",
        {
            "students": students
        }
    )


# ==========================================================
# Report Centre
# ==========================================================

@login_required
def report_center(request):

    return render(
        request,
        "reports/report_center.html"
    )


# ==========================================================
# Profile
# ==========================================================

@login_required
def profile(request):

    return render(
        request,
        "registration/profile.html"
    )


# ==========================================================
# Change Password
# ==========================================================

@login_required
def change_password(request):

    if request.method == "POST":

        form = PasswordChangeForm(
            request.user,
            request.POST
        )

        if form.is_valid():

            user = form.save()

            update_session_auth_hash(
                request,
                user
            )

            messages.success(
                request,
                "Password changed successfully."
            )

            return redirect(
                "profile"
            )

    else:

        form = PasswordChangeForm(
            request.user
        )

    return render(
        request,
        "registration/change_password.html",
        {
            "form": form
        }
    )


# ==========================================================
# Room Report PDF
# ==========================================================

@login_required
def room_report_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="room_report.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    elements = []

    styles = getSampleStyleSheet()

    elements.append(
        Paragraph(
            "SSMR MENS PG",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "Room Vacancy Report",
            styles["Heading2"]
        )
    )

    elements.append(
        Spacer(
            1,
            15
        )
    )

    data = [[

        "Room No",
        "Room Type",
        "Sharing",
        "Total Beds",
        "Occupied",
        "Available",
        "Monthly Rent",

    ]]

    rooms = Room.objects.all().order_by(
        "room_number"
    )

    for room in rooms:

        total_beds = (
            room.total_beds or 0
        )

        available_beds = (
            room.available_beds or 0
        )

        occupied_beds = (
            total_beds -
            available_beds
        )

        monthly_rent = (
            f"Rs.{room.monthly_rent}"
            if room.monthly_rent is not None
            else "Rs.0"
        )

        data.append([

            str(room.room_number),

            str(room.room_type),

            str(room.sharing_type),

            str(total_beds),

            str(occupied_beds),

            str(available_beds),

            monthly_rent,

        ])

    table = Table(
        data,
        repeatRows=1
    )

    table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkblue
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.whitesmoke
                ]
            ),

        ])
    )

    elements.append(
        table
    )

    doc.build(
        elements
    )

    return response


# ==========================================================
# Room Report Excel
# ==========================================================

@login_required
def room_report_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Room Report"

    headers = [

        "Room No",
        "Room Type",
        "Sharing Type",
        "Total Beds",
        "Occupied Beds",
        "Available Beds",
        "Monthly Rent",

    ]

    for col_num, header in enumerate(
        headers,
        1
    ):

        sheet.cell(
            row=1,
            column=col_num
        ).value = header

    rooms = Room.objects.all().order_by(
        "room_number"
    )

    row_num = 2

    for room in rooms:

        total_beds = (
            room.total_beds or 0
        )

        available_beds = (
            room.available_beds or 0
        )

        occupied_beds = (
            total_beds -
            available_beds
        )

        monthly_rent = (

            float(
                room.monthly_rent
            )

            if room.monthly_rent is not None

            else 0

        )

        sheet.cell(
            row=row_num,
            column=1
        ).value = room.room_number

        sheet.cell(
            row=row_num,
            column=2
        ).value = room.room_type

        sheet.cell(
            row=row_num,
            column=3
        ).value = room.sharing_type

        sheet.cell(
            row=row_num,
            column=4
        ).value = total_beds

        sheet.cell(
            row=row_num,
            column=5
        ).value = occupied_beds

        sheet.cell(
            row=row_num,
            column=6
        ).value = available_beds

        sheet.cell(
            row=row_num,
            column=7
        ).value = monthly_rent

        row_num += 1

    # ------------------------------------------------------
    # Widths
    # ------------------------------------------------------

    sheet.column_dimensions[
        "A"
    ].width = 15

    sheet.column_dimensions[
        "B"
    ].width = 18

    sheet.column_dimensions[
        "C"
    ].width = 18

    sheet.column_dimensions[
        "D"
    ].width = 15

    sheet.column_dimensions[
        "E"
    ].width = 18

    sheet.column_dimensions[
        "F"
    ].width = 18

    sheet.column_dimensions[
        "G"
    ].width = 18

    sheet.freeze_panes = "A2"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="room_report.xlsx"'
    )

    workbook.save(
        response
    )

    return response


# ==========================================================
# Room Print Report
# ==========================================================

@login_required
def room_report_print(request):

    rooms = Room.objects.all().order_by(
        "room_number"
    )

    return render(
        request,
        "rooms/room_report_print.html",
        {
            "rooms": rooms,
        }
    )


# ==========================================================
# Bed Report PDF
# ==========================================================

@login_required
def bed_report_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="bed_report.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    elements = []

    styles = getSampleStyleSheet()

    elements.append(
        Paragraph(
            "SSMR MENS PG",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "Bed Vacancy Report",
            styles["Heading2"]
        )
    )

    elements.append(
        Spacer(
            1,
            15
        )
    )

    data = [[

        "Room",
        "Bed",
        "Status",

    ]]

    beds = Bed.objects.select_related(
        "room"
    ).all().order_by(
        "room__room_number",
        "bed_number"
    )

    for bed in beds:

        room_number = "-"

        if bed.room:

            room_number = str(
                bed.room.room_number
            )

        bed_number = (

            str(bed.bed_number)

            if bed.bed_number

            else "-"

        )

        status = (

            str(bed.status)

            if bed.status

            else "-"

        )

        data.append([

            room_number,

            bed_number,

            status,

        ])

    table = Table(
        data,
        repeatRows=1
    )

    table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkblue
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                10
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                10
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black
            ),

            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.whitesmoke
                ]
            ),

        ])
    )

    elements.append(
        table
    )

    doc.build(
        elements
    )

    return response


# ==========================================================
# Bed Report Excel
# ==========================================================

@login_required
def bed_report_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Beds"

    sheet.append([
        "Room",
        "Bed",
        "Status",
    ])

    beds = Bed.objects.select_related(
        "room"
    ).all().order_by(
        "room__room_number",
        "bed_number"
    )

    for bed in beds:

        room_number = ""

        if bed.room:

            room_number = (
                bed.room.room_number
            )

        bed_number = (

            bed.bed_number

            if bed.bed_number

            else ""

        )

        status = (

            bed.status

            if bed.status

            else ""

        )

        sheet.append([
            room_number,
            bed_number,
            status,
        ])

    sheet.column_dimensions[
        "A"
    ].width = 20

    sheet.column_dimensions[
        "B"
    ].width = 20

    sheet.column_dimensions[
        "C"
    ].width = 20

    sheet.freeze_panes = "A2"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="bed_report.xlsx"'
    )

    workbook.save(
        response
    )

    return response


# ==========================================================
# Bed Report Print
# ==========================================================

@login_required
def bed_report_print(request):

    beds = Bed.objects.select_related(
        "room"
    ).all().order_by(
        "room__room_number",
        "bed_number"
    )

    return render(
        request,
        "beds/bed_report_print.html",
        {
            "beds": beds
        }
    )